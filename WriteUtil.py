import openpyxl
from openpyxl.styles.fonts import Font

'''
Write the data into a xlsx
'''
class WriteExcel:
    def __init__(self, fname):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.fname = fname
    
    '''
    Write data and save xlsx
    '''
    def write_excel(self, data):
        i, j = 1, 1
        for k,lst in data:
            self.sheet.cell(row=i, column=j).value = k
            # setup the font color
            if lst[1]:
                font_color = Font(color=lst[1])
                self.sheet.cell(row=i, column=j).font = font_color
            
            self.sheet.cell(row=i, column=j+1).value = lst[0]
            i += 1
        self.wb.save(self.fname)
