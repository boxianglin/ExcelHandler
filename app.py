import openpyxl
from ReadUtil import ReadStore
from WriteUtil import WriteExcel

# -------------------------------------- user APIs --------------------------------------------# 
def read_write_sorted():
    wb = openpyxl.load_workbook("Survey_snowballing.xlsx")
    title_year_col = [('A', 'B'), ('E', 'F')]
    
    
    rs = ReadStore(wb, title_year_col)
    rs.read_and_store()
    sorted_data = rs.data_year_sorted()
    
    we = WriteExcel('Sorted.xlsx')
    we.write_excel(sorted_data)
    

def merge_two_excel():
    title_year_col = [('A', 'B')]
    
    wb1 = openpyxl.load_workbook("Citations_Sorted.xlsx")
    rs1 = ReadStore(wb1, title_year_col)
    rs1.read_and_store()
    data1 = rs1.storage
    
    
    wb2 = openpyxl.load_workbook("SurveryAll_sorted.xlsx")
    rs2 = ReadStore(wb2, title_year_col)
    rs2.read_and_store()
    data2 = rs2.storage
    
    data2.update(data1)
    sorted_data = sorted(data2.items(), key=lambda x:x[1][0], reverse=True)
    
    we = WriteExcel('Merged.xlsx')
    we.write_excel(sorted_data)
    

if __name__ == '__main__':
    #read_write_sorted()
    merge_two_excel()
    
    
